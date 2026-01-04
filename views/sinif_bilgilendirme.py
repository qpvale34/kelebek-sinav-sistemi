"""
Sınıflara Göre Sınav Yerleri Bilgilendirme Ekranı
Öğrencilerin sınav/ salon dağılımlarını sınıf bazında gösterir ve yazdırır
"""

import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from assets.styles import (
    KelebekTheme,
    create_card_frame,
    configure_standard_button,
    show_message,
    ScrollableFrame
)
from assets.layout import setup_responsive_window
from controllers.database_manager import get_db
from utils import format_sira_label


class SinifBilgilendirmeView:
    """Sınıf bazlı sınav yerleri bilgilendirme ekranı"""

    def __init__(self, window, parent):
        self.window = window
        self.parent = parent
        self.db = get_db()

        self.class_data = {}
        self.selected_class = None

        self.window.title(f"{KelebekTheme.ICON_INFO} Sınıflara Göre Sınav Yerleri")
        self.window.geometry("1300x750")
        self.window.config(bg=KelebekTheme.BG_LIGHT)
        setup_responsive_window(self.window)

        self.setup_ui()
        self.load_data()

    # --------------------------- UI ---------------------------

    def setup_ui(self):
        header = tk.Frame(self.window, bg=KelebekTheme.ACCENT, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text=f"{KelebekTheme.ICON_INFO} HER SINIFA  SINAV YERLERİ BİLGİSİNİ Excel OLARAK KAYDET",
            font=(KelebekTheme.FONT_FAMILY, 20, "bold"),
            fg=KelebekTheme.TEXT_WHITE,
            bg=KelebekTheme.ACCENT
        ).pack(pady=15)

        # Ana içerik (Split View)
        content_frame = tk.Frame(self.window, bg=KelebekTheme.BG_LIGHT)
        content_frame.pack(fill="both", expand=True, padx=15, pady=15)

        # Sol panel - sınıf listesi (Sabit)
        left_panel = tk.Frame(content_frame, bg=KelebekTheme.BG_LIGHT, width=320)
        left_panel.pack(side="left", fill="y", padx=(0, 15))
        left_panel.pack_propagate(False)

        left_card = create_card_frame(left_panel, "Sınıf Listesi", KelebekTheme.ICON_STUDENT)
        left_card.pack(fill="both", expand=True)
        self.create_class_panel(left_card.content)

        # Sağ panel - detay (Esnek)
        right_panel = tk.Frame(content_frame, bg=KelebekTheme.BG_LIGHT)
        right_panel.pack(side="right", fill="both", expand=True)

        right_card = create_card_frame(right_panel, "Sınıf Detayları", KelebekTheme.ICON_INFO)
        right_card.pack(fill="both", expand=True)
        self.create_detail_panel(right_card.content)

    def create_class_panel(self, parent):
        search_frame = tk.Frame(parent, bg=KelebekTheme.BG_WHITE)
        search_frame.pack(fill="x", padx=10, pady=(10, 0))

        tk.Label(
            search_frame,
            text="Sınıf seviyesini seçin:",
            font=(KelebekTheme.FONT_FAMILY, 10, "bold"),
            bg=KelebekTheme.BG_WHITE
        ).pack(anchor="w")

        list_frame = tk.Frame(parent, bg=KelebekTheme.BG_WHITE)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.class_listbox = tk.Listbox(
            list_frame,
            height=20,
            font=(KelebekTheme.FONT_FAMILY, 11),
            activestyle="dotbox",
            exportselection=False
        )
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.class_listbox.yview)
        self.class_listbox.configure(yscrollcommand=scrollbar.set)

        self.class_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.class_listbox.bind("<<ListboxSelect>>", self.on_class_selected)

        btn_frame = tk.Frame(parent, bg=KelebekTheme.BG_WHITE)
        btn_frame.pack(fill="x", padx=10, pady=(0, 10))

        btn_refresh = tk.Button(btn_frame, command=self.load_data)
        configure_standard_button(btn_refresh, "secondary", "↻ Yenile")
        btn_refresh.pack(fill="x")

        self.class_info_label = tk.Label(
            parent,
            text="Sınıf seçilmedi",
            font=(KelebekTheme.FONT_FAMILY, 10),
            fg=KelebekTheme.TEXT_MUTED,
            bg=KelebekTheme.BG_WHITE,
            justify="left",
            wraplength=250
        )
        self.class_info_label.pack(fill="x", padx=10, pady=(0, 10))

    def create_detail_panel(self, parent):
        self.detail_info_label = tk.Label(
            parent,
            text="Sınıf seçilmedi.",
            font=(KelebekTheme.FONT_FAMILY, 11, "bold"),
            fg=KelebekTheme.TEXT_DARK,
            bg=KelebekTheme.BG_WHITE,
            anchor="w",
            justify="left"
        )
        self.detail_info_label.pack(fill="x", padx=15, pady=(10, 5))

        tree_frame = tk.Frame(parent, bg=KelebekTheme.BG_WHITE)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=10)

        columns = ("Ad", "Soyad", "Şube", "Sınav", "Ders", "Tarih", "Saat", "Salon", "Sıra")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=18)
        widths = [140, 140, 80, 220, 200, 110, 80, 120, 60]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)

        scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")

        btn_frame = tk.Frame(parent, bg=KelebekTheme.BG_WHITE)
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))

        btn_class = tk.Button(btn_frame, command=self.export_selected_class)
        configure_standard_button(btn_class, "secondary", "🖨️ Seçili Sınıfı Excel Olarak Kaydet")
        btn_class.pack(side="left", padx=5)

        btn_all = tk.Button(btn_frame, command=self.export_all_classes)
        configure_standard_button(btn_all, "success", "🖨️ HER SINIFA  SINAV YERLERİ BİLGİSİNİ Excel OLARAK KAYDET\n sınavdan önce sınıflara dağıt veya kat panolarına as \n herkes sınav yerini bilsin.")
        btn_all.pack(side="left", padx=5)

    # --------------------------- Data ---------------------------

    def load_data(self):
        self.class_data.clear()
        self.class_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.detail_info_label.config(text="Veriler yükleniyor...")
        self.class_info_label.config(text="Sınıf seçilmedi")

        try:
            sinavlar = self.db.harmanlanmis_sinavlar()
        except Exception as e:
            show_message(self.window, f"Sınav verileri alınamadı: {e}", "error")
            return

        if not sinavlar:
            self.detail_info_label.config(text="Harmanlanmış sınav bulunamadı.")
            return

        for sinav in sinavlar:
            try:
                yerlesimler = self.db.yerlesim_getir(sinav['id'])
            except Exception:
                continue
            if not yerlesimler:
                continue

            gozetmen_map = self._fetch_gozetmen_map(sinav['id'])

            for yer in yerlesimler:
                sinif_key = f"{yer['sinif']}. Sınıf - {yer['sube']} Şube"
                record = {
                    'ad': yer['ad'],
                    'soyad': yer['soyad'],
                    'sinif': yer['sinif'],
                    'sube': yer['sube'],
                    'salon': yer['salon_adi'],
                    'sira_no': yer['sira_no'],
                    'sira_label': format_sira_label(yer['sira_no']),
                    'sinav_adi': sinav['sinav_adi'],
                    'ders_adi': sinav['ders_adi'],
                    'tarih': sinav.get('sinav_tarihi') or '-',
                    'saat': sinav.get('sinav_saati') or '-',
                    'gozetmen': gozetmen_map.get(yer['salon_id'], "")
                }
                self.class_data.setdefault(sinif_key, []).append(record)

        if not self.class_data:
            self.detail_info_label.config(text="Öğrenci yerleşimi bulunamadı.")
            return

        for class_name in sorted(self.class_data.keys(), key=self._class_sort_key):
            count = len(self.class_data[class_name])
            self.class_listbox.insert(tk.END, f"{class_name} ({count})")

        self.detail_info_label.config(text="Sınıf seçimi yaparak detayları görüntüleyebilirsiniz.")
        if self.class_listbox.size() > 0:
            self.class_listbox.select_set(0)
            self.on_class_selected()

    def _fetch_gozetmen_map(self, sinav_id):
        try:
            atamalar = self.db.gozetmen_atamalari_listele(sinav_id)
        except AttributeError:
            return {}

        temp = {}
        for atama in atamalar:
            gorev = "Asıl" if atama['gorev_turu'] == 'asil' else "Yedek"
            temp.setdefault(atama['salon_id'], []).append(f"{atama['ad']} {atama['soyad']} ({gorev})")
        return {sid: ", ".join(names) for sid, names in temp.items()}

    # --------------------------- Events ---------------------------

    def on_class_selected(self, event=None):
        selection = self.class_listbox.curselection()
        if not selection:
            return
        display_text = self.class_listbox.get(selection[0])
        class_name = display_text.split("(")[0].strip()
        self.selected_class = class_name

        students = self.class_data.get(class_name, [])
        self.class_info_label.config(
            text=f"{class_name} için {len(students)} öğrencinin sınav yerleri listelenecek."
        )
        self.populate_students(class_name)

    def _class_sort_key(self, text):
        seviye = 0
        sube = ""
        try:
            left, _ = text.split("Sınıf", 1)
            seviye = int(left.strip().replace(".", ""))
            if '-' in text:
                sube = text.split('-')[-1].strip().split()[0]
        except Exception:
            pass
        return (seviye, sube)

    def populate_students(self, class_name):
        self.tree.delete(*self.tree.get_children())
        students = self.class_data.get(class_name, [])
        if not students:
            self.detail_info_label.config(text=f"{class_name} için kayıt bulunamadı.")
            return

        self.detail_info_label.config(
            text=f"{class_name} | Öğrenci: {len(students)}"
        )

        for record in sorted(students, key=lambda x: (x['tarih'], x['saat'], x['salon'], x['sira_no'])):
            self.tree.insert("", "end", values=(
                record['ad'],
                record['soyad'],
                record['sube'],
                record['sinav_adi'],
                record['ders_adi'],
                record['tarih'],
                record['saat'],
                record['salon'],
                record['sira_label']
            ))

    # --------------------------- Export ---------------------------

    def export_selected_class(self):
        if not self.selected_class:
            show_message(self.window, "Lütfen önce bir sınıf seçin!", "warning")
            return

        students = self.class_data.get(self.selected_class, [])
        if not students:
            show_message(self.window, "Bu sınıf için veri bulunamadı!", "warning")
            return

        base_dir = filedialog.askdirectory(title="Sınıf yerleşim dosyasının kaydedileceği klasörü seçin")
        if not base_dir:
            return

        folder_name = f"{self._safe_name(self.selected_class)}_{datetime.now().strftime('%Y%m%d_%H%M')}"
        target_dir = os.path.join(base_dir, folder_name)
        os.makedirs(target_dir, exist_ok=True)

        file_name = f"sinav_yerleri_{self._safe_name(self.selected_class)}.xlsx"
        file_path = os.path.join(target_dir, file_name)

        if self._write_excel(file_path, self.selected_class, students):
            show_message(self.window, f"📁 Dosya oluşturuldu:\n{target_dir}", "success")

    def export_all_classes(self):
        if not self.class_data:
            show_message(self.window, "Yazdırılacak sınıf bulunamadı!", "warning")
            return

        base_dir = filedialog.askdirectory(title="Tüm sınıf listelerinin kaydedileceği klasörü seçin")
        if not base_dir:
            return

        folder = os.path.join(base_dir, f"Kim_nerede_girecek_panolara_As_veya_siniflara_gonder{datetime.now().strftime('%Y%m%d_%H%M')}")
        os.makedirs(folder, exist_ok=True)

        saved = 0
        for class_name, students in self.class_data.items():
            if not students:
                continue
            safe = self._safe_name(class_name)
            file_name = f"sinav_yerleri_{safe}.xlsx"
            path = os.path.join(folder, file_name)
            if self._write_excel(path, class_name, students):
                saved += 1

        if saved:
            show_message(self.window, f"📁 {saved} sınıf için dosya oluşturuldu:\n{folder}", "success")
        else:
            show_message(self.window, "Herhangi bir dosya oluşturulamadı!", "warning")

    def _write_excel(self, path, class_name, students):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = class_name.replace(" ", "_")[:31]

            ws.merge_cells("A1:F1")
            ws["A1"] = f"{class_name} - Sınav Yerleri"
            ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
            ws["A1"].alignment = Alignment(horizontal="center")
            ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

            header_row = 3
            
            # Sınav sütunu da kaldırıldı
            headers = ["Ad", "Soyad", "Ders", "Salon", "Sıra"]
            for idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            current_row = header_row + 1
            for record in sorted(students, key=lambda x: (x['tarih'], x['saat'], x['salon'], x['sira_no'])):
                ws.cell(row=current_row, column=1, value=record['ad'])
                ws.cell(row=current_row, column=2, value=record['soyad'])
                
                ws.cell(row=current_row, column=3, value=record['ders_adi'])
                ws.cell(row=current_row, column=4, value=record['salon'])
                ws.cell(row=current_row, column=5, value=record['sira_label'])
                current_row += 1

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 8

            wb.save(path)
            return True
        except Exception as e:
            show_message(self.window, f"Dosya yazılırken hata oluştu:\n{e}", "error")
            return False

    def _safe_name(self, text):
        return "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in text).strip("_")


# Alias
ClassPlacementView = SinifBilgilendirmeView
