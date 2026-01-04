import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Parent dizini path'e ekle
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from assets.styles import KelebekTheme, configure_standard_button, show_message

class VisualSeatingPlanWindow(tk.Toplevel):
    """Görsel Renkli Sınıf Oturma Düzeni Penceresi (Gelişmiş)"""
    
    # Sınıf seviyelerine göre renkler
    CLASS_COLORS = {
        "5": "#3498db",   # Mavi
        "6": "#2ecc71",   # Yeşil
        "7": "#e67e22",   # Turuncu
        "8": "#9b59b6",   # Mor
        "9": "#f1c40f",   # Sarı
        "10": "#e74c3c",  # Kırmızı
        "11": "#95a5a6",  # Gri
        "12": "#795548",  # Kahverengi
        "Hazırlık": "#1abc9c" # Turkuaz
    }
    
    # Excel renk kodları (Hex - # kaldırılarak)
    EXCEL_COLORS = {k: v.replace("#", "") for k, v in CLASS_COLORS.items()}
    EXCEL_COLORS["Diğer"] = "BDC3C7" # Gri
    
    def __init__(self, parent, yerlesim_data, sinav_adi=""):
        super().__init__(parent)
        self.title(f"🎨 Görsel Sınıf Oturma Düzeni - {sinav_adi}")
        self.yerlesim_data = yerlesim_data
        self.sinav_adi = sinav_adi
        
        # State
        self.filtered_class = tk.StringVar(value="Tümü")
        self.all_classes = sorted(list(self.CLASS_COLORS.keys()), key=lambda x: int(x) if x.isdigit() else 99)
        
        # Tam ekran başlat
        self.state('zoomed')
        self.configure(bg=KelebekTheme.BG_LIGHT)
        
        self._build_ui()
        
    def _build_ui(self):
        # Sol Panel (Filtre ve İşlemler)
        left_panel = tk.Frame(self, bg=KelebekTheme.BG_WHITE, width=250)
        left_panel.pack(side="left", fill="y", padx=(0, 2))
        left_panel.pack_propagate(False)
        
        # Başlık Bölümü
        tk.Label(
            left_panel,
            text=f"🎨\nGörsel\nOturma Düzeni",
            font=(KelebekTheme.FONT_FAMILY, 16, "bold"),
            fg=KelebekTheme.PRIMARY,
            bg=KelebekTheme.BG_WHITE
        ).pack(pady=20)
        
        # Filtre Kartı
        filter_card = tk.LabelFrame(left_panel, text="Sınıf Filtresi", bg=KelebekTheme.BG_WHITE, fg=KelebekTheme.TEXT_DARK)
        filter_card.pack(fill="x", padx=10, pady=10)
        
        # Tümü Seçeneği
        rb_all = tk.Radiobutton(
            filter_card, 
            text="Tüm Sınıflar", 
            variable=self.filtered_class, 
            value="Tümü",
            bg=KelebekTheme.BG_WHITE,
            command=self._refresh_canvas
        )
        rb_all.pack(anchor="w", padx=5, pady=2)
        
        # Sınıflar
        for cls in self.all_classes:
            color = self.CLASS_COLORS.get(cls)
            frame = tk.Frame(filter_card, bg=KelebekTheme.BG_WHITE)
            frame.pack(fill="x", pady=1)
            
            rb = tk.Radiobutton(
                frame,
                text=f"{cls}. Sınıf",
                variable=self.filtered_class,
                value=cls,
                bg=KelebekTheme.BG_WHITE,
                command=self._refresh_canvas
            )
            rb.pack(side="left", padx=5)
            
            # Renk göstergesi
            tk.Label(frame, bg=color, width=2).pack(side="right", padx=10)
            
        # Kaydet Butonları
        tk.Label(left_panel, text="Dışa Aktar", font=("Arial", 10, "bold"), bg=KelebekTheme.BG_WHITE).pack(pady=(20, 5))
        
        # Sadece Excel butonu (HTML kaldırıldı)
        btn_save_excel = tk.Button(left_panel, command=self._export_to_excel)
        configure_standard_button(btn_save_excel, "primary", "📊 Excel Olarak Kaydet")
        btn_save_excel.pack(fill="x", padx=10, pady=5)
        
        tk.Label(
            left_panel,
            text="3 Blok x 5 Sıra (S Düzeni) formatında yazdırılabilir Excel çıktısı üretir.",
            font=("Arial", 8),
            fg=KelebekTheme.TEXT_MUTED,
            bg=KelebekTheme.BG_WHITE,
            wraplength=230,
            justify="center"
        ).pack(padx=10, pady=5)

        # Sağ Panel (Canvas)
        self.right_panel = tk.Frame(self, bg=KelebekTheme.BG_LIGHT)
        self.right_panel.pack(side="right", fill="both", expand=True)
        
        self.canvas = tk.Canvas(self.right_panel, bg=KelebekTheme.BG_LIGHT)
        self.scrollbar = ttk.Scrollbar(self.right_panel, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=KelebekTheme.BG_LIGHT)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Canvas genişliğini frame genişliğine eşitle (responsive grid için)
        self.canvas.bind('<Configure>', self._on_canvas_configure)
        
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Mousewheel
        self.canvas.bind_all("<MouseWheel>", lambda e: self.canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
        # İlk çizim
        self._refresh_canvas()
        
    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)
        
    def _refresh_canvas(self):
        # Temizle ve yeniden çiz
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self._draw_salons(self.scrollable_frame)
        
    def _draw_salons(self, parent):
        # Salonları grupla
        salons = {}
        for yer in self.yerlesim_data:
            salon_adi = yer['salon_adi']
            if salon_adi not in salons:
                salons[salon_adi] = []
            salons[salon_adi].append(yer)
            
        sorted_salon_names = sorted(salons.keys())
        
        container = tk.Frame(parent, bg=KelebekTheme.BG_LIGHT)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Grid ayarları
        row = 0
        col = 0
        max_col = 3
        
        for i in range(max_col):
             container.grid_columnconfigure(i, weight=1)
        
        target_class = self.filtered_class.get()
        
        for salon_adi in sorted_salon_names:
            ogrenciler = salons[salon_adi]
            self._create_salon_card(container, salon_adi, ogrenciler, row, col, target_class)
            
            col += 1
            if col >= max_col:
                col = 0
                row += 1
                
    def _create_salon_card(self, parent, salon_adi, ogrenciler, row, col, target_class):
        card = tk.Frame(parent, bg="white", bd=1, relief="solid")
        card.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
        
        # Salon Başlığı
        tk.Label(
            card,
            text=salon_adi,
            font=(KelebekTheme.FONT_FAMILY, 11, "bold"),
            bg="#f0f0f0",
            fg=KelebekTheme.TEXT_DARK,
            pady=5
        ).pack(fill="x")
        
        # Oturma düzeni alanı
        center_frame = tk.Frame(card, bg="white")
        center_frame.pack(expand=True, padx=10, pady=10)
        
        seating_frame = tk.Frame(center_frame, bg="white")
        seating_frame.pack()
        
        ogrenciler.sort(key=lambda x: int(x['sira_no']) if str(x['sira_no']).isdigit() else 999)
        
        # En büyük sıra numarasını bul
        max_sira = max([int(x['sira_no']) for x in ogrenciler if str(x['sira_no']).isdigit()], default=0)
        
        for ogrenci in ogrenciler:
            try:
                sira_no = int(ogrenci['sira_no'])
                r = (sira_no - 1) // 2
                c = (sira_no - 1) % 2
                real_col = 0 if c == 0 else 2
                
                # Renk belirleme
                sinif_seviyesi = "".join(filter(str.isdigit, str(ogrenci['sinif'])))
                if not sinif_seviyesi:
                    if "Hz" in str(ogrenci['sinif']) or "Haz" in str(ogrenci['sinif']):
                         cls_key = "Hazırlık"
                    else:
                         cls_key = "?"
                else:
                    cls_key = sinif_seviyesi
                
                base_color = self.CLASS_COLORS.get(cls_key, "#bdc3c7")
                
                # Filtre kontrolü
                if target_class != "Tümü" and cls_key != target_class:
                    bg_color = "#ecf0f1"
                    fg_color = "#bdc3c7"
                else:
                    bg_color = base_color
                    fg_color = "white"
                
                # Öğrenci Kutusu
                box = tk.Frame(seating_frame, bg=bg_color, width=70, height=55, bd=1, relief="raised")
                box.grid(row=r, column=real_col, padx=2, pady=2)
                box.pack_propagate(False)
                
                tk.Label(
                    box,
                    text=str(sira_no),
                    font=("Arial", 7, "bold"),
                    bg=bg_color,
                    fg=fg_color
                ).pack(anchor="nw", padx=1)
                
                if target_class == "Tümü" or cls_key == target_class:
                    tk.Label(
                        box,
                        text=f"{ogrenci.get('ad', '')}\n{ogrenci.get('soyad', '')}",
                        font=("Arial", 6),
                        bg=bg_color,
                        fg=fg_color,
                        wraplength=65
                    ).pack(expand=True)
                    
                    tk.Label(
                        box,
                        text=f"{ogrenci.get('sinif', '')}-{ogrenci.get('sube', '')}",
                        font=("Arial", 6, "bold"),
                        bg=bg_color,
                        fg=fg_color
                    ).pack(side="bottom", pady=1)
                
            except ValueError:
                continue

        # Koridor
        total_rows = (max_sira + 1) // 2
        tk.Label(seating_frame, text="K\nO\nR\nİ\nD\nO\nR", fg="#eee", font=("Arial", 8)).grid(row=0, column=1, rowspan=max(1, total_rows+1), padx=8)

    def _get_class_key(self, sinif_str):
        sinif_seviyesi = "".join(filter(str.isdigit, str(sinif_str)))
        if not sinif_seviyesi:
             if "Hz" in str(sinif_str) or "Haz" in str(sinif_str):
                 return "Hazırlık"
        return sinif_seviyesi if sinif_seviyesi else "Diğer"

    def _export_to_excel(self):
        """3x5 S Düzeninde Excel Çıktısı Al (A4 Uyumlu, Büyütülmüş ve Ortalanmış)"""
        file_path = filedialog.asksaveasfilename(
            title="Excel Olarak Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Sinav_Oturma_Duzeni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            # Varsayılan sheet'i sil
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Salonları grupla
            salons = {}
            for yer in self.yerlesim_data:
                s = yer['salon_adi']
                if s not in salons: salons[s] = []
                salons[s].append(yer)
            
            # Stiller
            border_style = Border(
                left=Side(style='medium'), right=Side(style='medium'), 
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            red_font = Font(bold=True, size=14, color="FF0000")
            
            for salon_adi, ogrenciler in salons.items():
                ws = wb.create_sheet(title=salon_adi[:30])
                
                # Sayfa Yapısı - A4 Yatay ve Sayfayı TAM Kapla
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToHeight = 1
                ws.page_setup.fitToWidth = 1
                ws.page_setup.horizontalCentered = True  # Yatay ortalama
                
                # Kenar boşluklarını minimize et (sayfayı maksimum kaplamak için)
                ws.page_margins.left = 0.2
                ws.page_margins.right = 0.2
                ws.page_margins.top = 0.2
                ws.page_margins.bottom = 0.2
                ws.page_margins.header = 0
                ws.page_margins.footer = 0
                
                # Sütun Genişlikleri (A4 YATAY SAYFAYI KAPLAMAK İÇİN BÜYÜK)
                # A4 Yatay: ~297mm genişlik, ~210mm yükseklik
                # Bloklar: A-B (Left), D-E (Mid), G-H (Right)
                # Koridorlar: C, F -> Dar
                std_width = 28  # MAKSİMUM genişlik
                cor_width = 6   # Koridor genişliği
                
                ws.column_dimensions['A'].width = std_width
                ws.column_dimensions['B'].width = std_width
                ws.column_dimensions['C'].width = cor_width
                ws.column_dimensions['D'].width = std_width
                ws.column_dimensions['E'].width = std_width
                ws.column_dimensions['F'].width = cor_width
                ws.column_dimensions['G'].width = std_width
                ws.column_dimensions['H'].width = std_width
                
                # 1. Başlık (Büyük)
                ws.merge_cells('A1:H1')
                ws['A1'] = f"{salon_adi} - SINAV OTURMA PLANI"
                ws['A1'].font = Font(bold=True, size=24)  # Daha büyük başlık
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 50
                
                # 2. Üst Satır: KAPI (Sol), TAHTA (Orta), ÖĞRETMEN MASASI (Sağ)
                ws.row_dimensions[2].height = 35
                
                # KAPI -> A2-B2 (En Sol, Kırmızı)
                ws.merge_cells('A2:B2')
                ws['A2'] = "KAPI"
                ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
                ws['A2'].font = red_font
                ws['A2'].fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                
                # TAHTA -> D2-E2 (Orta, Kırmızı)
                ws.merge_cells('D2:E2')
                ws['D2'] = "TAHTA"
                ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D2'].font = red_font
                ws['D2'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                
                # ÖĞRETMEN MASASI -> G2-H2 (Sağ)
                ws.merge_cells('G2:H2')
                ws['G2'] = "ÖĞRETMEN MASASI"
                ws['G2'].alignment = Alignment(horizontal='center', vertical='center')
                ws['G2'].font = Font(bold=True, size=12)
                ws['G2'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Koridor başlığı satırları
                ws.merge_cells('C3:C7')
                ws['C3'] = "K\nO\nR\nİ\nD\nO\nR"
                ws['C3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws['C3'].font = Font(size=10, color="666666", bold=True)
                ws['C3'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                ws.merge_cells('F3:F7')
                ws['F3'] = "K\nO\nR\nİ\nD\nO\nR"
                ws['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws['F3'].font = Font(size=10, color="666666", bold=True)
                ws['F3'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                # Sıralar Başlangıç: Row 3
                base_row_start = 3
                
                # Satır Yükseklikleri (ÖĞRENCİ MASALARI - SAYFA KAPLAMAK İÇİN BÜYÜK)
                # A4 yatay yükseklik ~210mm, header+kapi satırları ~30mm, kalan ~180mm / 5 satır = ~36mm = ~100 point
                for r in range(base_row_start, base_row_start + 5):
                    ws.row_dimensions[r].height = 100  # Maksimum yükseklik
                    
                # Öğrenci verilerini map'e al
                ogrenci_map = {}
                for ogr in ogrenciler:
                    try:
                        sn = int(ogr.get('sira_no', 0))
                        ogrenci_map[sn] = ogr
                    except:
                        pass
                
                # 30 sıralık grid oluştur (BOŞ dahil)
                for sira_no in range(1, 31):
                    block_idx = (sira_no - 1) // 10
                    local_idx = (sira_no - 1) % 10
                    row_in_block = local_idx // 2
                    side = local_idx % 2
                    
                    final_row = 0
                    final_col_idx = 0
                    
                    if block_idx == 0: # Sol (1-10) Front->Back
                         final_row = base_row_start + row_in_block
                         final_col_idx = 1 if side == 0 else 2
                         
                    elif block_idx == 1: # Orta (11-20) Back->Front
                         final_row = (base_row_start + 4) - row_in_block
                         final_col_idx = 4 if side == 0 else 5
                         
                    elif block_idx == 2: # Sağ (21-30) Front->Back
                         final_row = base_row_start + row_in_block
                         final_col_idx = 7 if side == 0 else 8
                    
                    col_letter = get_column_letter(final_col_idx)
                    cell = ws[f"{col_letter}{final_row}"]
                    
                    ogr = ogrenci_map.get(sira_no)
                    
                    if ogr:
                        ad = ogr.get('ad', '')
                        soyad = ogr.get('soyad', '')
                        sinif = ogr.get('sinif', '')
                        sube = ogr.get('sube', '')
                        cell.value = f"{sira_no}\n{ad} {soyad}\n{sinif}-{sube}"
                        cell.alignment = center_align
                        cell.border = border_style
                        
                        cls_key = self._get_class_key(sinif)
                        hex_color = self.EXCEL_COLORS.get(cls_key, "BDC3C7")
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.font = Font(size=11, color="FFFFFF", bold=True)  # Büyütüldü (9 -> 11)
                    else:
                        # Boş sıra
                        cell.value = f"{sira_no}\n\nBOŞ"
                        cell.alignment = center_align
                        cell.border = border_style
                        cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                        cell.font = Font(size=11, color="999999", italic=True)
                        
            wb.save(file_path)
            show_message(self, f"✅ Excel kaydedildi:\n{file_path}", "success")
            os.startfile(file_path)
            
        except Exception as e:
            show_message(self, f"Excel kayıt hatası: {e}", "error")

