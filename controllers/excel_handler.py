"""
Kelebek Sınav Sistemi - Excel Handler
Excel dosyalarından veri okuma ve yazma işlemleri
"""

from typing import List, Dict, Tuple, Optional
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError(
        "pandas modülü bulunamadı. 'pip install -r requirements.txt' veya 'pip install pandas' komutunu çalıştırın."
    ) from exc

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ModuleNotFoundError as exc:
    raise ModuleNotFoundError(
        "openpyxl modülü bulunamadı. 'pip install -r requirements.txt' ya da 'pip install openpyxl' komutunu çalıştırın."
    ) from exc

from utils import MIN_SINIF, MAX_SINIF, SINIF_SEVIYELERI, format_sira_label, sinif_ismi_gecerli_mi


class ExcelHandler:
    """Excel işlemlerini yöneten sınıf"""
    
    # Excel şablon kolonları
    OGRENCI_KOLONLARI = ['ad', 'soyad', 'sinif', 'sube', 'tc_no']
    GOZETMEN_KOLONLARI = ['ad', 'soyad', 'email', 'telefon']
    
    @staticmethod
    def ogrenci_sablonu_olustur(dosya_yolu: str) -> bool:
        """
        Öğrenci ekleme için örnek Excel şablonu oluştur
        """
        try:
            # Örnek veri - yeni sınıf seviyeleri ile
            ornek_data = {
                'ad': ['Ahmet', 'Ayşe', 'Mehmet', 'Zeynep', 'Ali', 'Fatma', 'Can'],
                'soyad': ['YILMAZ', 'KAYA', 'DEMİR', 'ŞAHİN', 'ÇELIK', 'ÖZTÜRK', 'ARSLAN'],
                'sinif': ['5', '8', '11sayisal', '12sozel', 'lisehazirlikingilizce', 'ortaokulhazirlikarapca', '10'],
                'sube': ['A', 'A', 'B', 'B', 'C', 'A', 'D'],
                'tc_no': ['12345678901', '12345678902', '12345678903', '12345678904', '12345678905', '12345678906', '12345678907']
            }
            
            df = pd.DataFrame(ornek_data)
            
            # Excel'e yaz
            with pd.ExcelWriter(dosya_yolu, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Öğrenciler', index=False)
                
                # Worksheet'i al ve formatla
                workbook = writer.book
                worksheet = writer.sheets['Öğrenciler']
                
                # Başlık formatı
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Kolon genişlikleri - sınıf kolonu geniş (string sınıflar için)
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 25
                worksheet.column_dimensions['D'].width = 10
                worksheet.column_dimensions['E'].width = 15
            
            return True
        except Exception as e:
            print(f"❌ Şablon oluşturma hatası: {e}")
            return False
    
    @staticmethod
    def gozetmen_sablonu_olustur(dosya_yolu: str) -> bool:
        """
        Gözetmen ekleme için örnek Excel şablonu oluştur
        """
        try:
            ornek_data = {
                'ad': ['Ahmet', 'Ayşe', 'Mehmet'],
                'soyad': ['ÖZTÜRK', 'YILDIRIM', 'ARSLAN'],
                'email': ['ahmet@okul.com', 'ayse@okul.com', 'mehmet@okul.com'],
                'telefon': ['555-1234', '555-5678', '555-9012']
            }
            
            df = pd.DataFrame(ornek_data)
            
            with pd.ExcelWriter(dosya_yolu, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Gözetmenler', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Gözetmenler']
                
                header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 25
                worksheet.column_dimensions['D'].width = 15
            
            return True
        except Exception as e:
            print(f"❌ Şablon oluşturma hatası: {e}")
            return False
    
    @staticmethod
    def ogrenci_oku(dosya_yolu: str) -> Tuple[List[Dict], List[str]]:
        """
        Excel dosyasından öğrenci listesi oku
        Returns: (başarılı_kayıtlar, hata_mesajları)
        """
        hatalar = []
        basarili_kayitlar = []
        
        try:
            # Excel dosyasını oku
            df = pd.read_excel(dosya_yolu)
            
            # Kolon kontrolü
            gerekli_kolonlar = set(ExcelHandler.OGRENCI_KOLONLARI)
            mevcut_kolonlar = set(df.columns.str.lower().str.strip())
            
            eksik_kolonlar = gerekli_kolonlar - mevcut_kolonlar
            if eksik_kolonlar:
                hatalar.append(f"❌ Eksik kolonlar: {', '.join(eksik_kolonlar)}")
                return [], hatalar
            
            # Kolon isimlerini normalize et
            df.columns = df.columns.str.lower().str.strip()
            
            # Her satırı işle
            for idx, row in df.iterrows():
                satir_no = idx + 2  # Excel satır numarası (başlık hariç)
                
                try:
                    # Boş satır kontrolü
                    if pd.isna(row['ad']) or pd.isna(row['soyad']):
                        continue
                    
                    # Veri temizleme ve validasyon
                    sinif_deger = str(row['sinif']).strip()  # Olduğu gibi sakla
                    # Sayısal sınıf kontrolü (5, 6, 7, vb.)
                    if sinif_deger.isdigit():
                        sinif_deger = sinif_deger  # String olarak tut
                    
                    ogrenci_data = {
                        'ad': str(row['ad']).strip(),
                        'soyad': str(row['soyad']).strip(),
                        'sinif': sinif_deger,  # String olarak sakla
                        'sube': str(row['sube']).strip(),  # Olduğu gibi sakla (esnek)
                        'tc_no': str(row['tc_no']).strip() if pd.notna(row['tc_no']) else None
                    }
                    
                    # Sınıf kontrolü (esnek validasyon - int() dönüşümü YOK)
                    # 1. Önce sabit listede var mı kontrol et
                    # 2. Yoksa regex ile geçerli karakterler mi kontrol et
                    sinif_str = ogrenci_data['sinif']
                    
                    if sinif_str not in SINIF_SEVIYELERI:
                        # Sabit listede yok, regex ile kontrol et
                        if not sinif_ismi_gecerli_mi(sinif_str):
                            hatalar.append(
                                f"⚠️ Satır {satir_no}: Geçersiz sınıf karakterleri ({sinif_str}). "
                                f"Sadece harf, sayı, Türkçe karakterler ve - _ kullanılabilir."
                            )
                            continue
                    
                    # Sınıf değeri string olarak kalıyor, dönüşüm yok
                    
                    # TC No kontrolü (varsa)
                    if ogrenci_data['tc_no']:
                        tc_temiz = ogrenci_data['tc_no'].replace('-', '').replace(' ', '')
                        if len(tc_temiz) != 11 or not tc_temiz.isdigit():
                            hatalar.append(f"⚠️ Satır {satir_no}: Geçersiz TC No ({ogrenci_data['tc_no']})")
                            ogrenci_data['tc_no'] = None
                        else:
                            ogrenci_data['tc_no'] = tc_temiz
                    
                    basarili_kayitlar.append(ogrenci_data)
                    
                except Exception as e:
                    hatalar.append(f"❌ Satır {satir_no}: {str(e)}")
                    continue
            
            if not basarili_kayitlar:
                hatalar.append("❌ Hiç geçerli kayıt bulunamadı!")
            
            return basarili_kayitlar, hatalar
            
        except FileNotFoundError:
            hatalar.append(f"❌ Dosya bulunamadı: {dosya_yolu}")
            return [], hatalar
        except Exception as e:
            hatalar.append(f"❌ Dosya okuma hatası: {str(e)}")
            return [], hatalar
    
    @staticmethod
    def gozetmen_oku(dosya_yolu: str) -> Tuple[List[Dict], List[str]]:
        """
        Excel dosyasından gözetmen listesi oku
        Returns: (başarılı_kayıtlar, hata_mesajları)
        """
        hatalar = []
        basarili_kayitlar = []
        
        try:
            df = pd.read_excel(dosya_yolu)
            
            # Kolon kontrolü
            gerekli_kolonlar = {'ad', 'soyad'}  # email ve telefon opsiyonel
            mevcut_kolonlar = set(df.columns.str.lower().str.strip())
            
            eksik_kolonlar = gerekli_kolonlar - mevcut_kolonlar
            if eksik_kolonlar:
                hatalar.append(f"❌ Eksik kolonlar: {', '.join(eksik_kolonlar)}")
                return [], hatalar
            
            df.columns = df.columns.str.lower().str.strip()
            
            for idx, row in df.iterrows():
                satir_no = idx + 2
                
                try:
                    if pd.isna(row['ad']) or pd.isna(row['soyad']):
                        continue
                    
                    gozetmen_data = {
                        'ad': str(row['ad']).strip(),
                        'soyad': str(row['soyad']).strip(),
                        'email': str(row['email']).strip() if 'email' in df.columns and pd.notna(row['email']) else None,
                        'telefon': str(row['telefon']).strip() if 'telefon' in df.columns and pd.notna(row['telefon']) else None
                    }
                    
                    basarili_kayitlar.append(gozetmen_data)
                    
                except Exception as e:
                    hatalar.append(f"❌ Satır {satir_no}: {str(e)}")
                    continue
            
            if not basarili_kayitlar:
                hatalar.append("❌ Hiç geçerli kayıt bulunamadı!")
            
            return basarili_kayitlar, hatalar
            
        except FileNotFoundError:
            hatalar.append(f"❌ Dosya bulunamadı: {dosya_yolu}")
            return [], hatalar
        except Exception as e:
            hatalar.append(f"❌ Dosya okuma hatası: {str(e)}")
            return [], hatalar
    
    @staticmethod
    def yerlesim_yazdir(dosya_yolu: str, sinav_bilgi: Dict, yerlesim_data: List[Dict],
                        gozetmen_data: Dict = None) -> bool:
        """
        Sınav yerleşimini Excel'e yazdır (sadece öğrenci ve salon bilgileri)
        
        Args:
            dosya_yolu: Kaydedilecek dosya yolu
            sinav_bilgi: Sınav bilgileri (ders_adi, tarih, saat, sinav_adi vb.)
            yerlesim_data: Yerleşim verileri (ogrenci, salon, sira bilgileri)
            gozetmen_data: Salon bazlı gözetmen bilgileri {salon_id: [gozetmen_listesi]}
        """
        try:
            # Salonlara göre grupla
            salonlar = {}
            salon_gozetmen_map = {}
            for yer in yerlesim_data:
                salon_adi = yer['salon_adi']
                if salon_adi not in salonlar:
                    salonlar[salon_adi] = []
                salonlar[salon_adi].append(yer)
                # Gözetmen bilgisini salon_id bazlı kaydet
                if 'gozetmenler' in yer and yer['gozetmenler']:
                    salon_gozetmen_map[salon_adi] = yer['gozetmenler']
            
            # Harici gözetmen_data parametresi varsa birleştir
            if gozetmen_data:
                for salon_id, gozetmenler in gozetmen_data.items():
                    # salon_id'yi salon_adi'ne dönüştür gerekirse
                    if isinstance(gozetmenler, str):
                        salon_gozetmen_map[str(salon_id)] = gozetmenler
            
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Varsayılan sheet'i sil
            
            # Sınav adı bilgisi
            sinav_adi = sinav_bilgi.get('sinav_adi', sinav_bilgi.get('ders_adi', 'Sınav'))
            
            # Her salon için ayrı sheet
            for salon_adi, ogrenciler in sorted(salonlar.items()):
                ws = wb.create_sheet(title=salon_adi[:31])  # Excel sheet isim limiti
                
                # Başlık bilgileri
                ws.merge_cells('A1:F1')
                ws['A1'] = f"SINAV YERLEŞİM LİSTESİ"
                ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30
                
                # Salon bilgisi
                ws['A2'] = f"Salon: {salon_adi} | Öğrenci Sayısı: {len(ogrenciler)}"
                ws['A2'].font = Font(bold=True, size=11)
                
                header_row = 4
                
                # Tablo başlıkları (YOKLAMA eklendi)
                headers = ['Sıra No', 'Ad', 'Soyad', 'Sınıf', 'Şube', 'YOKLAMA']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Öğrenci verileri
                data_start_row = header_row + 1
                last_data_row = data_start_row
                for idx, ogr in enumerate(sorted(ogrenciler, key=lambda x: x.get('sira_no', 0)), start=data_start_row):
                    # Güvenli alan erişimi
                    ws.cell(row=idx, column=1, value=format_sira_label(ogr.get('sira_no', '')))
                    ws.cell(row=idx, column=2, value=ogr.get('ad', ''))
                    ws.cell(row=idx, column=3, value=ogr.get('soyad', ''))
                    ws.cell(row=idx, column=4, value=ogr.get('sinif', ''))
                    ws.cell(row=idx, column=5, value=ogr.get('sube', ''))
                    ws.cell(row=idx, column=6, value='')  # YOKLAMA sütunu (boş)
                    last_data_row = idx
                    
                    # Zebra efekti
                    if idx % 2 == 0:
                        fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        for col in range(1, 7):  # 6 sütun oldu
                            ws.cell(row=idx, column=col).fill = fill
                
                # Kenarlıklar
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=header_row, max_row=last_data_row, min_col=1, max_col=6):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # ============ GÖZETMENLERİ BÖLÜMÜ ============
                # 2 satır boşluk bırak
                gozetmen_section_start = last_data_row + 3
                
                # Gözetmenler başlık satırı (adı, soyadı, tarih, imza)
                gozetmen_headers = ['Gözetmenler', 'adı', 'soyadı', 'tarih', 'imza']
                header_cols = [1, 2, 3, 4, 5]  # A, B, C, D, E sütunları
                
                # Başlık hücrelerini ayarla
                for col, header in zip(header_cols, gozetmen_headers):
                    cell = ws.cell(row=gozetmen_section_start, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Gözetmen satırları (numaralı boş satırlar)
                # Bu salon için gözetmen bilgisi var mı kontrol et
                salon_gozetmenler = salon_gozetmen_map.get(salon_adi, "")
                
                # Gözetmen isimlerini parse et
                gozetmen_listesi = []
                if salon_gozetmenler:
                    # "Ad Soyad (Asıl), Ad2 Soyad2 (Yedek)" formatından ayır
                    parts = [g.strip() for g in salon_gozetmenler.split(',')]
                    for part in parts:
                        # "Ad Soyad (Görev)" formatını ayır
                        if '(' in part:
                            isim_kisim = part.split('(')[0].strip()
                        else:
                            isim_kisim = part.strip()
                        
                        if isim_kisim:
                            isim_parcalar = isim_kisim.split()
                            if len(isim_parcalar) >= 2:
                                ad = ' '.join(isim_parcalar[:-1])
                                soyad = isim_parcalar[-1]
                            else:
                                ad = isim_kisim
                                soyad = ''
                            gozetmen_listesi.append({'ad': ad, 'soyad': soyad})
                
                # En az 3 gözetmen satırı oluştur (numaralı)
                gozetmen_satir_sayisi = max(3, len(gozetmen_listesi))
                
                for i in range(gozetmen_satir_sayisi):
                    row_num = gozetmen_section_start + 1 + i
                    
                    # 1. sütun: numara
                    ws.cell(row=row_num, column=1, value=i + 1)
                    
                    if i < len(gozetmen_listesi):
                        # Mevcut gözetmen bilgisi
                        ws.cell(row=row_num, column=2, value=gozetmen_listesi[i]['ad'])
                        ws.cell(row=row_num, column=3, value=gozetmen_listesi[i]['soyad'])
                    else:
                        # Boş satır (elle doldurulmak üzere)
                        ws.cell(row=row_num, column=2, value='')
                        ws.cell(row=row_num, column=3, value='')
                    
                    ws.cell(row=row_num, column=4, value='')  # Tarih alanı (elle doldurulacak)
                    ws.cell(row=row_num, column=5, value='')  # İmza alanı
                    
                    # Kenarlık ekle
                    for col in range(1, 6):
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Kolon genişlikleri
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 10
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 15  # YOKLAMA sütunu
            
            # Kaydet
            wb.save(dosya_yolu)
            return True
            
        except Exception as e:
            print(f"❌ Excel yazma hatası: {e}")
            return False
    
    @staticmethod
    def yoklama_formu_olustur(dosya_yolu: str, sinav_bilgi: Dict, 
                              yerlesim_data: List[Dict]) -> bool:
        """
        Yoklama formu oluştur (gözetmenler için)
        """
        try:
            salonlar = {}
            for yer in yerlesim_data:
                salon_adi = yer['salon_adi']
                if salon_adi not in salonlar:
                    salonlar[salon_adi] = []
                salonlar[salon_adi].append(yer)
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            for salon_adi, ogrenciler in sorted(salonlar.items()):
                ws = wb.create_sheet(title=f"{salon_adi} Yoklama"[:31])
                
                # Başlık
                ws.merge_cells('A1:F1')
                ws['A1'] = f"YOKLAMA FORMU - {salon_adi}"
                ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                
                ws['A3'] = f"Ders: {sinav_bilgi['ders_adi']} | Tarih: {sinav_bilgi['tarih']} | Saat: {sinav_bilgi['saat']}"
                ws['A3'].font = Font(bold=True)
                
                # Gözetmen bilgisi
                ws['A5'] = "Gözetmen Adı Soyadı:"
                ws['A5'].font = Font(bold=True)
                ws['C5'] = "İmza:"
                ws['C5'].font = Font(bold=True)
                
                # Tablo başlıkları
                headers = ['Sıra', 'Ad', 'Soyad', 'Sınıf/Şube', 'Var', 'İmza']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=7, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Öğrenciler
                for idx, ogr in enumerate(sorted(ogrenciler, key=lambda x: x['sira_no']), start=8):
                    ws.cell(row=idx, column=1, value=format_sira_label(ogr['sira_no']))
                    ws.cell(row=idx, column=2, value=ogr['ad'])
                    ws.cell(row=idx, column=3, value=ogr['soyad'])
                    ws.cell(row=idx, column=4, value=f"{ogr['sinif']}/{ogr['sube']}")
                    ws.cell(row=idx, column=5, value="☐")  # Checkbox
                    ws.cell(row=idx, column=6, value="")  # İmza alanı
                
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 6
                ws.column_dimensions['F'].width = 20
            
            wb.save(dosya_yolu)
            return True
            
        except Exception as e:
            print(f"❌ Yoklama formu oluşturma hatası: {e}")
            return False

if __name__ == "__main__":
    # Test
    handler = ExcelHandler()
    
    print("🧪 Excel Handler Test\n")
    
    # Şablon oluştur
    if handler.ogrenci_sablonu_olustur("test_ogrenci_sablon.xlsx"):
        print("✅ Öğrenci şablonu oluşturuldu: test_ogrenci_sablon.xlsx")
    
    if handler.gozetmen_sablonu_olustur("test_gozetmen_sablon.xlsx"):
        print("✅ Gözetmen şablonu oluşturuldu: test_gozetmen_sablon.xlsx")
